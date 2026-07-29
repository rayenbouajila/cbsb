"""
excel_export.py

Genere un classeur Excel (.xlsx) a partir d'une liste de factures dont
CHACUNE peut avoir un jeu de champs different (extraction dynamique).

Principe : on calcule l'UNION de tous les champs reellement presents parmi
les factures a exporter, dans un ordre stable (champs comptables connus
d'abord, puis champs "extra" specifiques a certaines factures). Une facture
qui n'a pas un champ donne laisse simplement la cellule correspondante vide -
il n'y a pas de structure figee imposee a toutes les lignes.

Quand c'est possible sans ambiguite, une formule Excel reconstitue une valeur
manquante a partir des autres colonnes de la meme ligne (ex: TVA = HT * taux/100
si la facture a HT + taux mais pas de montant TVA extrait directement).
"""

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------- Palette corporate bleu / gris ----------
COLOR_PRIMARY = "1F3B57"
COLOR_PRIMARY_LIGHT = "2E5A87"
COLOR_ACCENT = "4A90D9"
COLOR_ACCENT_2 = "6FA8DC"
COLOR_GRAY_LIGHT = "F2F4F6"
COLOR_GRAY_BORDER = "D9DEE3"
COLOR_WHITE = "FFFFFF"
COLOR_TEXT = "1A1A1A"

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
FONT_KPI_LABEL = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_KPI_VALUE = Font(name="Calibri", size=18, bold=True, color=COLOR_WHITE)
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=COLOR_PRIMARY)
FONT_BODY = Font(name="Calibri", size=10, color=COLOR_TEXT)

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_GRAY_BORDER),
    right=Side(style="thin", color=COLOR_GRAY_BORDER),
    top=Side(style="thin", color=COLOR_GRAY_BORDER),
    bottom=Side(style="thin", color=COLOR_GRAY_BORDER),
)

# Champs comptables "connus" : ordre d'affichage prioritaire + libelle fixe.
# Doit rester coherent avec CANONICAL_LABELS dans invoice_extraction.py.
CANONICAL_LABELS = {
    "numero_facture": "N° Facture",
    "date_facture": "Date",
    "fournisseur": "Fournisseur",
    "categorie": "Catégorie",
    "montant_ht": "Montant HT",
    "taux_tva": "Taux TVA (%)",
    "montant_tva": "Montant TVA",
    "montant_ttc": "Montant TTC",
}
CANONICAL_ORDER = list(CANONICAL_LABELS.keys())
MONETARY_KEYS = ("montant_ht", "montant_tva", "montant_ttc")


@dataclass
class InvoiceExport:
    """Une facture a exporter. `fields` est la liste dynamique telle que
    stockee dans invoice_extracted_data.extracted_fields (JSONB) :
    [{"key": "montant_ht", "label": "Montant HT", "value": 1000.0, "value_type": "number"}, ...]
    Deux InvoiceExport peuvent avoir des `fields` totalement differents."""
    filename: str
    client_name: Optional[str] = None
    fields: list = field(default_factory=list)


def _coerce_date(value):
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value
    return value


def _compute_union_columns(invoices: list[InvoiceExport]) -> list[dict]:
    """Calcule l'ensemble des colonnes a afficher : union des `key` presents
    dans au moins une facture, champs comptables connus en premier (dans un
    ordre fixe), puis les champs 'extra' dans l'ordre de premiere apparition."""
    first_seen: dict[str, dict] = {}
    for inv in invoices:
        for f in inv.fields:
            if f["key"] not in first_seen:
                first_seen[f["key"]] = f

    canonical_cols = [first_seen[k] for k in CANONICAL_ORDER if k in first_seen]
    extra_cols = [f for k, f in first_seen.items() if k not in CANONICAL_ORDER]
    return canonical_cols + extra_cols


def _style_header_row(ws: Worksheet, row: int, start_col: int, end_col: int, fill_color: str = COLOR_PRIMARY):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autofit_columns(ws: Worksheet, min_width: int = 10, max_width: int = 42):
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_letter = get_column_letter(cell.column)
            length = len(str(cell.value))
            widths[col_letter] = max(widths.get(col_letter, 0), length)
    for col_letter, length in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 4, max_width))


def _build_detail_sheet(wb: Workbook, invoices: list[InvoiceExport], columns: list[dict]) -> Worksheet:
    ws = wb.create_sheet("Detail Factures")

    headers = ["Client", "Fichier"] + [c["label"] for c in columns]
    ws.append(headers)
    _style_header_row(ws, row=1, start_col=1, end_col=len(headers))
    ws.freeze_panes = "A2"

    fill_alt = PatternFill(start_color=COLOR_GRAY_LIGHT, end_color=COLOR_GRAY_LIGHT, fill_type="solid")

    # Colonne fixe (1-based) de chaque champ, pour construire les formules
    col_index = {c["key"]: 3 + i for i, c in enumerate(columns)}
    col_letters = {key: get_column_letter(idx) for key, idx in col_index.items()}

    first_data_row = 2
    for i, inv in enumerate(invoices):
        r = first_data_row + i
        fields_by_key = {f["key"]: f for f in inv.fields}

        ws.cell(row=r, column=1, value=inv.client_name or "-")
        ws.cell(row=r, column=2, value=inv.filename or "-")

        for col in columns:
            key = col["key"]
            cell_col = col_index[key]
            f = fields_by_key.get(key)

            if f is not None:
                # Valeur reellement extraite pour CETTE facture
                value = f["value"]
                cell = ws.cell(row=r, column=cell_col)
                if f["value_type"] == "number":
                    try:
                        cell.value = float(value)
                        cell.number_format = "0.00" if key == "taux_tva" else '#,##0.000 "DT"'
                    except (TypeError, ValueError):
                        cell.value = str(value)
                elif f["value_type"] == "date":
                    cell.value = _coerce_date(value)
                    cell.number_format = "DD/MM/YYYY"
                else:
                    cell.value = str(value)

            elif key == "montant_tva" and "montant_ht" in fields_by_key and "taux_tva" in fields_by_key:
                # Pas de TVA extraite directement, mais HT + taux presents -> formule
                ht_l, taux_l = col_letters["montant_ht"], col_letters["taux_tva"]
                cell = ws.cell(row=r, column=cell_col, value=f"={ht_l}{r}*({taux_l}{r}/100)")
                cell.number_format = '#,##0.000 "DT"'

            elif key == "montant_ttc" and "montant_ht" in fields_by_key:
                # Pas de TTC extrait directement, mais HT present -> formule
                ht_l = col_letters["montant_ht"]
                if "montant_tva" in fields_by_key or ("taux_tva" in fields_by_key and "montant_tva" == key):
                    tva_l = col_letters.get("montant_tva")
                    formula = f"={ht_l}{r}+{tva_l}{r}" if tva_l else f"={ht_l}{r}"
                else:
                    formula = f"={ht_l}{r}"
                cell = ws.cell(row=r, column=cell_col, value=formula)
                cell.number_format = '#,##0.000 "DT"'

            # sinon : la facture n'a tout simplement pas ce champ -> cellule vide

        # Style de ligne (alternance + bordures) sur toutes les colonnes
        row_fill = fill_alt if i % 2 == 1 else None
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_num)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left" if col_num <= 2 else "right")
            if row_fill:
                cell.fill = row_fill

    last_data_row = first_data_row + len(invoices) - 1 if invoices else first_data_row
    total_row = last_data_row + 1

    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=2).font = Font(bold=True, color=COLOR_PRIMARY)

    for col in columns:
        if col["value_type"] != "number":
            continue
        cell_col = col_index[col["key"]]
        letter = get_column_letter(cell_col)
        formula = f"=SUM({letter}{first_data_row}:{letter}{last_data_row})" if invoices else "=0"
        cell = ws.cell(row=total_row, column=cell_col, value=formula)
        cell.number_format = "0.00" if col["key"] == "taux_tva" else '#,##0.000 "DT"'
        cell.font = Font(bold=True, color=COLOR_PRIMARY)
        cell.border = Border(top=Side(style="double", color=COLOR_PRIMARY))

    _autofit_columns(ws)
    ws.sheet_view.showGridLines = False

    return ws, col_index, total_row, first_data_row, last_data_row


def _build_synthese_sheet(
    wb: Workbook,
    invoices: list[InvoiceExport],
    columns: list[dict],
    col_index: dict,
    total_row: int,
    detail_sheet_name: str = "Detail Factures",
) -> Worksheet:
    ws = wb.create_sheet("Synthese", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H2")
    title_cell = ws["B2"]
    title_cell.value = "Synthese des Factures"
    title_cell.font = FONT_TITLE

    ws.merge_cells("B3:H3")
    subtitle_cell = ws["B3"]
    subtitle_cell.value = f"{len(invoices)} facture(s) selectionnee(s) - champs variables selon les documents sources"
    subtitle_cell.font = Font(size=10, italic=True, color="6B7280")

    # KPI uniquement pour les colonnes monetaires REELLEMENT presentes dans la selection
    present_monetary = [c for c in columns if c["key"] in MONETARY_KEYS]
    kpi_colors = {
        "montant_ht": COLOR_PRIMARY,
        "montant_tva": COLOR_PRIMARY_LIGHT,
        "montant_ttc": COLOR_ACCENT,
    }

    kpi_start_row = 5
    if not present_monetary:
        ws.merge_cells("B5:H6")
        note = ws["B5"]
        note.value = "Aucune donnee monetaire structuree n'a pu etre extraite pour cette selection."
        note.font = Font(italic=True, color="8A9A93")
        note.alignment = Alignment(horizontal="left", vertical="center")
    else:
        for i, col in enumerate(present_monetary):
            key = col["key"]
            col_start = 2 + i * 3
            col_end = col_start + 1
            color = kpi_colors.get(key, COLOR_ACCENT_2)
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            label_row = kpi_start_row
            ws.merge_cells(start_row=label_row, start_column=col_start, end_row=label_row, end_column=col_end)
            label_cell = ws.cell(row=label_row, column=col_start, value=CANONICAL_LABELS[key].upper())
            label_cell.font = FONT_KPI_LABEL
            label_cell.fill = fill
            label_cell.alignment = Alignment(horizontal="center", vertical="center")

            value_row = kpi_start_row + 1
            ws.merge_cells(start_row=value_row, start_column=col_start, end_row=value_row, end_column=col_end)
            letter = get_column_letter(col_index[key])
            value_cell = ws.cell(row=value_row, column=col_start, value=f"='{detail_sheet_name}'!{letter}{total_row}")
            value_cell.number_format = '#,##0.000 "DT"'
            value_cell.font = FONT_KPI_VALUE
            value_cell.fill = fill
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[label_row].height = 22
            ws.row_dimensions[value_row].height = 34

    # Recap
    recap_row = kpi_start_row + 4
    ws.cell(row=recap_row, column=2, value="Nombre de factures").font = Font(bold=True, size=10)
    ws.cell(row=recap_row, column=4, value=len(invoices))

    # Liste des colonnes presentes dans cette export (utile pour comprendre
    # pourquoi certaines factures ont des cases vides dans le detail)
    fields_row = recap_row + 2
    ws.cell(row=fields_row, column=2, value="Champs detectes dans cette selection :").font = Font(bold=True, size=10)
    for i, col in enumerate(columns):
        ws.cell(row=fields_row + 1 + i, column=2, value=f"• {col['label']}").font = Font(size=9.5, color="5A6E67")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 3

    return ws


def build_invoices_excel(invoices: list[InvoiceExport]) -> io.BytesIO:
    """Point d'entree principal. Prend une liste d'InvoiceExport (potentiellement
    avec des `fields` totalement differents d'une facture a l'autre), retourne
    un buffer BytesIO pret pour une reponse HTTP."""
    wb = Workbook()
    wb.remove(wb.active)

    columns = _compute_union_columns(invoices)

    _, col_index, total_row, _, _ = _build_detail_sheet(wb, invoices, columns)
    _build_synthese_sheet(wb, invoices, columns, col_index, total_row)

    wb.active = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer